import openpyxl
from tkinter import filedialog
from tkinter import messagebox
from threading import Thread
import sqlite3


class ImportData:
    def __init__(self, db_path: str) -> None:
        self.db_path = db_path

        # Excel File
        self.working_book = ""
        self.sheet_on_working_book = []

        # Sheet
        self.sheet_new_income = ""
        self.sheet_reactivation = ""
        self.sheet_suspension = ""
        self.sheet_discharge = ""
        self.sheet_workers = ""

    def setWorkBook(self) -> bool:
        filename = filedialog.askopenfilename()

        if filename == "":
            return False

        try:
            self.working_book = openpyxl.load_workbook(
                filename=filename,
                read_only=True,
                data_only=True,
                keep_links=False
            )
        except openpyxl.utils.exceptions.InvalidFileException:
            messagebox.showerror(
                "Error", "Extensión de archivo invalido debe de ser .xlsx")
            return False

    def setSheetsToWorkWith(self):
        self.sheet_on_working_book = self.working_book.sheetnames
        try:
            self.sheet_new_income = self.sheet_on_working_book[1]
            self.sheet_reactivation = self.sheet_on_working_book[2]
            self.sheet_suspension = self.sheet_on_working_book[3]
            self.sheet_discharge = self.sheet_on_working_book[4]
            self.sheet_workers = self.sheet_on_working_book[5]
        except IndexError:
            return False

    def setWorkersInDB(self):
        worker_sheet = self.working_book[self.sheet_workers]
        info_worker = []

        for row in worker_sheet.iter_rows(values_only=True, min_col=1, max_col=20, min_row=14):
            if row[0] == None:
                break

            row = list(row)
            row[0] = int(row[0])

            info_worker.append(row)

        with sqlite3.connect(self.db_path) as bd:
            cursor = bd.cursor()
            cursor.executemany(
                """INSERT INTO workers VALUES
                (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                info_worker
            )
            bd.commit()

    def setNewIncomeInDB(self):
        new_income_sheet = self.working_book[self.sheet_new_income]
        info_worker = []

        for row in new_income_sheet.iter_rows(values_only=True, min_col=1, max_col=21, min_row=12):
            if row[0] == None:
                break

            row = list(row)
            info_worker.append([row[2], row[20]])

        with sqlite3.connect(self.db_path) as bd:
            cursor = bd.cursor()
            cursor.executemany(
                """INSERT INTO newIncome VALUES(?,?)""",
                info_worker
            )
            bd.commit()

        return True

    def setReactivationInDB(self):
        reactivation_sheet = self.working_book[self.sheet_reactivation]
        info_worker = []

        for row in reactivation_sheet.iter_rows(values_only=True, min_col=1, max_col=22, min_row=12):
            if row[0] == None:
                break

            row = list(row)
            info_worker.append([row[2], row[21], row[8]])

        with sqlite3.connect(self.db_path) as bd:
            cursor = bd.cursor()
            cursor.executemany(
                """INSERT INTO reactivations VALUES(?,?,?)""",
                info_worker
            )
            bd.commit()

        return True

    def setDischargeInDB(self):
        discharge_sheet = self.working_book[self.sheet_discharge]
        info_worker = []

        for row in discharge_sheet.iter_rows(values_only=True, min_col=1, max_col=22, min_row=12):
            if row[0] == None:
                break

            row = list(row)
            info_worker.append([row[2], row[18], row[19], row[21]])

        with sqlite3.connect(self.db_path) as bd:
            cursor = bd.cursor()
            cursor.executemany(
                """INSERT INTO discharge VALUES(?,?,?,?)""",
                info_worker
            )
            bd.commit()

        return True

    def setSuspensionsInDB(self):
        suspension_sheet = self.working_book[self.sheet_suspension]
        info_worker = []

        for row in suspension_sheet.iter_rows(values_only=True, min_col=1, max_col=22, min_row=12):
            if row[0] == None:
                break

            row = list(row)
            info_worker.append([row[2], row[18], row[19], row[21]])

        with sqlite3.connect(self.db_path) as bd:
            cursor = bd.cursor()
            cursor.executemany(
                """INSERT INTO suspensions VALUES(?,?,?,?)""",
                info_worker
            )

            bd.commit()

    def operate(self):
        if self.setWorkBook() == False:
            return False

        if self.setSheetsToWorkWith() == False:
            return False

        threads = [
            Thread(target=self.setWorkersInDB),
            Thread(target=self.setNewIncomeInDB),
            Thread(target=self.setReactivationInDB),
            Thread(target=self.setDischargeInDB),
            Thread(target=self.setSuspensionsInDB)
        ]

        for thread in threads:
            thread.start()
            thread.join()

        return True
